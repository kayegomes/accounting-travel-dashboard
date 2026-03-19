#!/usr/bin/env python3.11
"""
Analisador Completo V3 - Painel Contábil
Versão COMPLETA com IA, todos os componentes e geração de planilha tratada + Orçado vs Realizado
✅ CORRIGIDO: Remoção de duplicatas por Passageiro+Data+Valor
✅ NOVO: Extração de valores orçados da aba RESUMO LOGÍSTICA
"""

import pandas as pd
import openpyxl
from datetime import datetime
import json
import os
from pathlib import Path
from openai import OpenAI
import traceback

class AnalisadorCompletoV3:
    """Classe principal para análise completa com IA"""
    
    PLATAFORMAS_PRIORIDADE = ["AMAZON", "COMBATE", "GLOBOPLAY", "PREMIERE", "SPORTV", "TV GLOBO"]
    
    def __init__(self, caminho_planilha):
        self.caminho_planilha = Path(caminho_planilha)
        self.wb = None
        self.dados_consolidados = {}
        self.client = None
        self.filtros = {}
        
        # Validar se arquivo existe
        if not self.caminho_planilha.exists():
            raise FileNotFoundError(f"❌ Arquivo não encontrado: {self.caminho_planilha}")
        
        # Inicializar cliente OpenAI se disponível
        try:
            api_key = os.getenv('OPENAI_API_KEY')
            if api_key:
                self.client = OpenAI(api_key=api_key)
                print("✅ Cliente IA inicializado")
            else:
                print("⚠️  OPENAI_API_KEY não configurada - continuando sem IA")
        except Exception as e:
            print(f"⚠️  Cliente IA não disponível - continuando sem IA: {e}")
        
    def carregar_planilha(self):
        """Carrega a planilha Excel e também lê filtros do próprio arquivo"""
        print("📂 Carregando planilha...")
        try:
            self.wb = openpyxl.load_workbook(str(self.caminho_planilha), data_only=True, read_only=True)
            print(f"✅ Planilha carregada: {len(self.wb.sheetnames)} abas")
            try:
                self.filtros = self._ler_filtros_da_planilha()
                print("🔎 Filtros lidos da planilha:", self.filtros)
            except Exception as e:
                print("⚠️ Não foi possível ler filtros da planilha automaticamente:", e)
                self.filtros = {}
        except Exception as e:
            raise Exception(f"❌ Erro ao carregar planilha: {e}")
    
    def _validar_aba(self, nome_aba):
        """Valida se uma aba existe na planilha"""
        if not self.wb:
            raise Exception("Planilha não foi carregada. Execute carregar_planilha() primeiro.")
        if nome_aba not in self.wb.sheetnames:
            raise Exception(f"❌ Aba '{nome_aba}' não encontrada. Abas disponíveis: {', '.join(self.wb.sheetnames)}")
        return True
    
    def _validar_colunas(self, df, colunas_esperadas, nome_componente):
        """Valida se as colunas esperadas existem no DataFrame"""
        colunas_faltantes = [col for col in colunas_esperadas if col not in df.columns]
        if colunas_faltantes:
            raise Exception(f"❌ Colunas faltantes em {nome_componente}: {', '.join(colunas_faltantes)}")
        return True
    
    def _ordenar_plataformas(self, df, coluna='Plataforma'):
        """Ordena DataFrame de plataformas de acordo com a prioridade definida"""
        if df is None or len(df) == 0 or coluna not in df.columns:
            return df
        
        df = df.copy()
        prioridade_map = {nome: idx for idx, nome in enumerate(self.PLATAFORMAS_PRIORIDADE)}
        df['_plat_norm'] = df[coluna].astype(str).str.strip().str.upper()
        df['_ordem_dashboard'] = df['_plat_norm'].map(prioridade_map)
        
        if df['_ordem_dashboard'].notna().any():
            df['_ordem_dashboard'] = df['_ordem_dashboard'].fillna(len(self.PLATAFORMAS_PRIORIDADE))
            df = df.sort_values(['_ordem_dashboard', next((c for c in ['Valor_Total', 'Valor', 'Valor_passagens'] if c in df.columns), '_ordem_dashboard')], ascending=[True, False])
        else:
            df = df.sort_values(next((c for c in ['Valor_Total', 'Valor', 'Valor_passagens'] if c in df.columns), coluna), ascending=False)
        
        return df.drop(columns=['_plat_norm', '_ordem_dashboard'], errors='ignore')

    def _ler_filtros_da_planilha(self):
        """Lê filtros da planilha Excel"""
        nome_aba = None
        for n in self.wb.sheetnames:
            if 'RESUMO' in n.upper() or 'FILTROS' in n.upper():
                nome_aba = n
                break
        if not nome_aba:
            nome_aba = self.wb.sheetnames[0]
        
        ws = self.wb[nome_aba]
        labels_map = {
            'ÁREA': 'AREA',
            'AREA': 'AREA',
            'PRODUTO MACRO': 'PRODUTO_MACRO',
            'PRODUTO': 'PRODUTO_MACRO',
            'NATUREZA': 'NATUREZA',
            'PLATAFORMA': 'PLATAFORMA',
            'MÊS': 'MES',
            'MES': 'MES',
            'ACUMULADO': 'MES'
        }
        filtros = {'AREA': None, 'PRODUTO_MACRO': None, 'NATUREZA': None, 'PLATAFORMA': None, 'OPERADOR_MES': None, 'MES': None}
        
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=5):
            for cell in row:
                try:
                    val = cell.value
                    if val is None:
                        continue
                    key = str(val).strip().upper()
                    if key in labels_map:
                        coluna_dest = labels_map[key]
                        right_col = cell.column + 1
                        candidate = ws.cell(row=cell.row, column=right_col).value
                        if candidate is not None:
                            filtros[coluna_dest] = candidate
                        else:
                            candidate2 = ws.cell(row=cell.row, column=right_col+1).value
                            if candidate2 is not None:
                                filtros[coluna_dest] = candidate2
                except Exception:
                    continue
        
        try:
            c24 = ws['C24'].value
            d24 = ws['D24'].value
            if c24 is not None:
                try:
                    filtros['MES'] = int(float(c24))
                except Exception:
                    try:
                        filtros['MES'] = int(''.join([c for c in str(c24) if c.isdigit()]))
                    except:
                        pass
            if d24 is not None:
                op = str(d24).strip()
                if op in ['=', '<=', '<', '>=']:
                    filtros['OPERADOR_MES'] = op
        except Exception:
            pass
        
        if filtros.get('OPERADOR_MES') is None:
            for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=5):
                for cell in row:
                    v = cell.value
                    if v is None:
                        continue
                    s = str(v).strip()
                    if s in ['<=', '=']:
                        filtros['OPERADOR_MES'] = s
                        break
                if filtros.get('OPERADOR_MES'):
                    break
        
        for k in ['AREA','PRODUTO_MACRO','NATUREZA','PLATAFORMA']:
            v = filtros.get(k)
            if v is None:
                continue
            sv = str(v).strip()
            if sv == '' or sv == '*' or sv.upper() == 'TODOS':
                filtros[k] = None
            else:
                filtros[k] = sv
        
        if filtros.get('MES') is not None:
            try:
                filtros['MES'] = int(float(filtros['MES']))
            except:
                filtros['MES'] = None
        
        if filtros.get('OPERADOR_MES') is None:
            filtros['OPERADOR_MES'] = '<=' if filtros.get('MES') is not None else None
        
        return filtros

    def extrair_passagens(self):
        """✅ CORRIGIDO: Extrai dados de passagens com remoção de duplicatas"""
        print("\n✈️  Processando PASSAGENS...")

        nome_aba = 'BasePassagens_New'
        self._validar_aba(nome_aba)

        try:
            df = pd.read_excel(str(self.caminho_planilha), sheet_name=nome_aba)
            print("👉 Colunas BasePassagens_New:", df.columns.tolist())
        except Exception as e:
            raise Exception(f"❌ Erro ao ler aba {nome_aba}: {e}")

        # Detectar colunas
        col_valor = next((c for c in df.columns if c.strip().upper() in ['VALOR AJUSTADO','VALOR_AJUSTADO','VALOR_AJUST', 'VALORAJUSTADO']), None)
        col_plata = next((c for c in df.columns if c.strip().upper() in ['PLATAFORMA','PLATFORMA','PLATAF']), None)
        col_nomeproj = next((c for c in df.columns if c.strip().upper() in ['NOME PROJETO','NOME_PROJETO','NOME_PROJ']), None)
        col_area = next((c for c in df.columns if c.strip().upper() in ['ÁREA','AREA']), None)
        col_natureza = next((c for c in df.columns if c.strip().upper() in ['NATUREZA']), None)
        col_mes_z = next((c for c in df.columns if c.strip().upper() in ['Z','MÊS','MES','MONTH','PERIODO']), None)
        col_prod_macro = next((c for c in df.columns if c.strip().upper() in ['PRODUTO MACRO','PRODUTO_MACRO','PRODUTO']), None)
        col_data = next((c for c in df.columns if c.strip().upper() in ['DATA','DATE']), None)
        col_passageiro = next((c for c in df.columns if c.strip().upper() in ['PASSAGEIRO','PASSAGEIROS','PESSOA']), None)

        # Renomear
        df = df.rename(columns={
            col_valor: 'VALOR_AJUSTADO',
            col_plata: 'PLATAFORMA',
            col_nomeproj: 'NOME_PROJETO',
            col_area: 'AREA',
            col_natureza: 'NATUREZA',
            col_mes_z: 'MES_Z',
            col_prod_macro: 'PRODUTO_MACRO',
            col_data: 'DATA',
            col_passageiro: 'PASSAGEIRO'
        })

        # Limpar
        df['VALOR_AJUSTADO'] = pd.to_numeric(df['VALOR_AJUSTADO'], errors='coerce')
        df['MES_Z'] = pd.to_numeric(df['MES_Z'], errors='coerce')

        # FILTRO ANO = 2025
        if 'DATA' in df.columns:
            df['DATA'] = pd.to_datetime(df['DATA'], errors='coerce')
            df = df[df['DATA'].dt.year == 2025]

        # Aplicar filtros da planilha
        filtros = self.filtros or {}
        df_f = df.copy()

        # ❌ REMOVIDO: Não filtrar Reemissões (Excel inclui todas)
        # O Excel NÃO remove reemissões, então não devemos remover também

        # FILTRO Área
        areas_filter = filtros.get('AREA')
        if isinstance(areas_filter, str) and ',' in areas_filter:
            areas_list = [a.strip().lower() for a in areas_filter.split(',') if a.strip() != '']
        elif isinstance(areas_filter, (list, tuple)):
            areas_list = [a.strip().lower() for a in areas_filter if str(a).strip() != '']
        elif areas_filter:
            areas_list = [str(areas_filter).strip().lower()]
        else:
            areas_list = None

        if areas_list:
            if 'AREA' in df_f.columns:
                df_f = df_f[df_f['AREA'].astype(str).str.strip().str.lower().isin(areas_list)]

        # FILTRO Produto Macro
        produto_macro = filtros.get('PRODUTO_MACRO')
        if produto_macro:
            if 'PRODUTO_MACRO' in df_f.columns:
                df_f = df_f[df_f['PRODUTO_MACRO'].astype(str).str.strip() == str(produto_macro).strip()]

        # FILTRO Natureza
        natureza = filtros.get('NATUREZA')
        if natureza:
            if 'NATUREZA' in df_f.columns:
                df_f = df_f[df_f['NATUREZA'].astype(str).str.strip() == str(natureza).strip()]

        # FILTRO Mês
        operador = filtros.get('OPERADOR_MES')
        mes = filtros.get('MES')
        if 'MES_Z' in df_f.columns and mes is not None and operador:
            if operador == "=":
                df_f = df_f[df_f['MES_Z'] == mes]
            elif operador == "<=":
                df_f = df_f[df_f['MES_Z'] <= mes]

        # FILTRO Plataforma
        plataforma = filtros.get('PLATAFORMA')
        if plataforma and 'PLATAFORMA' in df_f.columns:
            df_f = df_f[df_f['PLATAFORMA'].astype(str).str.strip() == str(plataforma).strip()]

        df_f = df_f[df_f['VALOR_AJUSTADO'].notna()]

        # ✅ FILTRO CRÍTICO: Apenas as 12 áreas que o Excel soma
        areas_excel = ['Elenco', 'Colaborador', 'Convidado', 'Gestão Eventos', 'Motorista',
                       'Influencer', 'Sup. eventos', 'Ed. Eventos', 'Produção de Eventos',
                       'Repórter', 'Repcine', 'Demais']
        
        if 'AREA' in df_f.columns:
            # Case-insensitive match para lidar com variações de maiúsculas/minúsculas
            df_f['_area_lower'] = df_f['AREA'].astype(str).str.strip().str.lower()
            areas_excel_lower = [a.lower() for a in areas_excel]
            df_f = df_f[df_f['_area_lower'].isin(areas_excel_lower)]
            df_f = df_f.drop(columns=['_area_lower'])
            print(f"   🔍 Filtradas {len(df_f)} registros nas {len(areas_excel)} áreas do Excel")

        # ✅ Remover valores negativos (ajustes/estornos)
        if 'VALOR_AJUSTADO' in df_f.columns:
            registros_negativos = (df_f['VALOR_AJUSTADO'] < 0).sum()
            if registros_negativos > 0:
                valor_negativo = df_f[df_f['VALOR_AJUSTADO'] < 0]['VALOR_AJUSTADO'].sum()
                df_f = df_f[df_f['VALOR_AJUSTADO'] >= 0]
                print(f"   🔧 Removidos {registros_negativos} valores negativos (R$ {valor_negativo:,.2f})")

        # ✅ SOLUÇÃO CRÍTICA: Remover duplicatas por Passageiro + Data + VALOR AJUSTADO
        if 'PASSAGEIRO' in df_f.columns and 'DATA' in df_f.columns and 'VALOR_AJUSTADO' in df_f.columns:
            registros_antes = len(df_f)
            df_f = df_f.drop_duplicates(subset=['PASSAGEIRO', 'DATA', 'VALOR_AJUSTADO'], keep='first')
            registros_removidos = registros_antes - len(df_f)
            if registros_removidos > 0:
                print(f"   🔧 Removidas {registros_removidos} duplicatas (Passageiro+Data+Valor)")

        if len(df_f) == 0:
            print("   ⚠️  Nenhum dado válido encontrado em Passagens após aplicar filtros")
            total = 0
            qtd = 0
        else:
            total = df_f['VALOR_AJUSTADO'].sum()
            qtd = len(df_f)

        # Agrupar
        if qtd > 0:
            por_plataforma = df_f.groupby('PLATAFORMA')['VALOR_AJUSTADO'].agg(['sum', 'count']).reset_index()
            por_plataforma.columns = ['Plataforma', 'Valor', 'Quantidade']
            por_plataforma = self._ordenar_plataformas(por_plataforma)

            por_campeonato = df_f.groupby('NOME_PROJETO')['VALOR_AJUSTADO'].agg(['sum', 'count']).reset_index()
            por_campeonato.columns = ['Campeonato', 'Valor', 'Quantidade']
            por_campeonato = por_campeonato.sort_values('Valor', ascending=False)

            por_grupo = df_f.groupby('AREA')['VALOR_AJUSTADO'].agg(['sum', 'count']).reset_index()
            por_grupo.columns = ['Grupo', 'Valor', 'Quantidade']
            por_grupo = por_grupo.sort_values('Valor', ascending=False)
        else:
            por_plataforma = pd.DataFrame(columns=['Plataforma', 'Valor', 'Quantidade'])
            por_campeonato = pd.DataFrame(columns=['Campeonato', 'Valor', 'Quantidade'])
            por_grupo = pd.DataFrame(columns=['Grupo', 'Valor', 'Quantidade'])

        self.dados_consolidados['passagens'] = {
            'total': total,
            'quantidade': qtd,
            'por_plataforma': por_plataforma,
            'por_campeonato': por_campeonato,
            'por_grupo': por_grupo,
            'dataframe': df_f
        }

        print(f"   Total: R$ {total:,.2f} | Qtd: {qtd:,}")

        
    def extrair_hospedagens(self):
        """✅ CORRIGIDO: Extrai dados de hospedagens com remoção de duplicatas"""
        print("\n🏨 Processando HOSPEDAGENS...")

        nome_aba = 'BaseHospedagens_New'
        self._validar_aba(nome_aba)

        try:
            df = pd.read_excel(str(self.caminho_planilha), sheet_name=nome_aba)
            print("👉 Colunas BaseHospedagens_New:", df.columns.tolist())
        except Exception as e:
            raise Exception(f"❌ Erro ao ler aba {nome_aba}: {e}")

        # Detectar colunas
        col_valor = next((c for c in df.columns if c.strip().upper() in ['TOTAL AJUSTADO','TOTAL_AJUSTADO','TOTAL AJUST']), None)
        col_plata = next((c for c in df.columns if c.strip().upper() in ['PLATAFORMA','PLATAFORMA','PLATAFORMA']), None)
        col_nomeproj = next((c for c in df.columns if c.strip().upper() in ['NOME PROJETO','NOME_PROJETO']), None)
        col_area = next((c for c in df.columns if c.strip().upper() in ['ÁREA','AREA']), None)
        col_natureza = next((c for c in df.columns if c.strip().upper() in ['NATUREZA']), None)
        col_mes = next((c for c in df.columns if c.strip().upper() in ['MÊS','MES','MONTH','PERIODO']), None)
        col_prod_macro = next((c for c in df.columns if c.strip().upper() in ['PLATAFORMA','PRODUTO MACRO','PRODUTO_MACRO','PRODUTO']), None)
        col_data = next((c for c in df.columns if c.strip().upper() in ['DATA','DATE','CHECK-IN','CHECKIN']), None)
        col_hospede = next((c for c in df.columns if c.strip().upper() in ['HÓSPEDE','HOSPEDE','PESSOA']), None)

        # Renomear
        rename_map = {}
        if col_valor: rename_map[col_valor] = 'TOTAL_AJUSTADO'
        if col_plata: rename_map[col_plata] = 'PLATAFORMA'
        if col_nomeproj: rename_map[col_nomeproj] = 'NOME_PROJETO'
        if col_area: rename_map[col_area] = 'AREA'
        if col_natureza: rename_map[col_natureza] = 'NATUREZA'
        if col_mes: rename_map[col_mes] = 'MES'
        if col_prod_macro: rename_map[col_prod_macro] = 'PRODUTO_MACRO'
        if col_data: rename_map[col_data] = 'DATA'
        if col_hospede: rename_map[col_hospede] = 'HOSPEDE'
        df = df.rename(columns=rename_map)

        # Limpar
        if 'TOTAL_AJUSTADO' in df.columns:
            df['TOTAL_AJUSTADO'] = pd.to_numeric(df['TOTAL_AJUSTADO'], errors='coerce')
        if 'MES' in df.columns:
            df['MES'] = pd.to_numeric(df['MES'], errors='coerce')

        # FILTRO ANO = 2025
        if 'DATA' in df.columns:
            df['DATA'] = pd.to_datetime(df['DATA'], errors='coerce')
            df = df[df['DATA'].dt.year == 2025]

        # Aplicar filtros
        filtros = self.filtros or {}
        df_f = df.copy()

        # ❌ REMOVIDO: Não filtrar Reemissões (Excel inclui todas)
        # O Excel NÃO remove reemissões, então não devemos remover também

        # FILTRO Área
        areas_filter = filtros.get('AREA')
        if isinstance(areas_filter, str) and ',' in areas_filter:
            areas_list = [a.strip().lower() for a in areas_filter.split(',') if a.strip() != '']
        elif isinstance(areas_filter, (list, tuple)):
            areas_list = [a.strip().lower() for a in areas_filter if str(a).strip() != '']
        elif areas_filter:
            areas_list = [str(areas_filter).strip().lower()]
        else:
            areas_list = None

        if areas_list:
            if 'AREA' in df_f.columns:
                df_f = df_f[df_f['AREA'].astype(str).str.strip().str.lower().isin(areas_list)]

        # FILTRO Produto Macro
        produto_macro = filtros.get('PRODUTO_MACRO')
        if produto_macro and 'PRODUTO_MACRO' in df_f.columns:
            df_f = df_f[df_f['PRODUTO_MACRO'].astype(str).str.strip() == str(produto_macro).strip()]

        # FILTRO Natureza
        natureza = filtros.get('NATUREZA')
        if natureza and 'NATUREZA' in df_f.columns:
            df_f = df_f[df_f['NATUREZA'].astype(str).str.strip() == str(natureza).strip()]

        # FILTRO Mês
        operador = filtros.get('OPERADOR_MES')
        mes = filtros.get('MES')
        if 'MES' in df_f.columns and mes is not None and operador:
            if operador == "=":
                df_f = df_f[df_f['MES'] == mes]
            elif operador == "<=":
                df_f = df_f[df_f['MES'] <= mes]

        # FILTRO Plataforma
        plataforma = filtros.get('PLATAFORMA')
        if plataforma and 'PLATAFORMA' in df_f.columns:
            df_f = df_f[df_f['PLATAFORMA'].astype(str).str.strip() == str(plataforma).strip()]

        # Excluir cancelados
        if 'AREA' in df_f.columns:
            df_f = df_f[df_f['AREA'].astype(str).str.strip().str.upper() != 'CANCELADO']

        if 'TOTAL_AJUSTADO' in df_f.columns:
            df_f = df_f[df_f['TOTAL_AJUSTADO'].notna()]

        # ✅ FILTRO CRÍTICO: Apenas as 12 áreas que o Excel soma
        areas_excel = ['Elenco', 'Colaborador', 'Convidado', 'Gestão Eventos', 'Motorista',
                       'Influencer', 'Sup. eventos', 'Ed. Eventos', 'Produção de Eventos',
                       'Repórter', 'Repcine', 'Demais']
        
        if 'AREA' in df_f.columns:
            # Case-insensitive match
            df_f['_area_lower'] = df_f['AREA'].astype(str).str.strip().str.lower()
            areas_excel_lower = [a.lower() for a in areas_excel]
            df_f = df_f[df_f['_area_lower'].isin(areas_excel_lower)]
            df_f = df_f.drop(columns=['_area_lower'])
            print(f"   🔍 Filtradas {len(df_f)} registros nas {len(areas_excel)} áreas do Excel")

        # ✅ Remover valores negativos (ajustes/estornos)
        if 'TOTAL_AJUSTADO' in df_f.columns:
            registros_negativos = (df_f['TOTAL_AJUSTADO'] < 0).sum()
            if registros_negativos > 0:
                valor_negativo = df_f[df_f['TOTAL_AJUSTADO'] < 0]['TOTAL_AJUSTADO'].sum()
                df_f = df_f[df_f['TOTAL_AJUSTADO'] >= 0]
                print(f"   🔧 Removidos {registros_negativos} valores negativos (R$ {valor_negativo:,.2f})")

        # ✅ SOLUÇÃO CRÍTICA: Remover duplicatas por Hóspede + Data + TOTAL_AJUSTADO
        if 'HOSPEDE' in df_f.columns and 'DATA' in df_f.columns and 'TOTAL_AJUSTADO' in df_f.columns:
            registros_antes = len(df_f)
            df_f = df_f.drop_duplicates(subset=['HOSPEDE', 'DATA', 'TOTAL_AJUSTADO'], keep='first')
            registros_removidos = registros_antes - len(df_f)
            if registros_removidos > 0:
                print(f"   🔧 Removidas {registros_removidos} duplicatas (Hóspede+Data+Valor)")

        if len(df_f) == 0:
            print("   ⚠️  Nenhum dado válido encontrado em Hospedagens após aplicar filtros")
            total = 0
            qtd = 0
        else:
            total = df_f['TOTAL_AJUSTADO'].sum()
            qtd = len(df_f)

        # Agrupar
        if qtd > 0 and 'PLATAFORMA' in df_f.columns and 'TOTAL_AJUSTADO' in df_f.columns:
            por_plataforma = df_f.groupby('PLATAFORMA')['TOTAL_AJUSTADO'].agg(['sum', 'count']).reset_index()
            por_plataforma.columns = ['Plataforma', 'Valor', 'Quantidade']
            por_plataforma = self._ordenar_plataformas(por_plataforma)

            por_campeonato = df_f.groupby('NOME_PROJETO')['TOTAL_AJUSTADO'].agg(['sum', 'count']).reset_index()
            por_campeonato.columns = ['Campeonato', 'Valor', 'Quantidade']
            por_campeonato = por_campeonato.sort_values('Valor', ascending=False)

            por_grupo = df_f.groupby('AREA')['TOTAL_AJUSTADO'].agg(['sum', 'count']).reset_index()
            por_grupo.columns = ['Grupo', 'Valor', 'Quantidade']
            por_grupo = por_grupo.sort_values('Valor', ascending=False)
        else:
            por_plataforma = pd.DataFrame(columns=['Plataforma', 'Valor', 'Quantidade'])
            por_campeonato = pd.DataFrame(columns=['Campeonato', 'Valor', 'Quantidade'])
            por_grupo = pd.DataFrame(columns=['Grupo', 'Valor', 'Quantidade'])

        self.dados_consolidados['hospedagens'] = {
            'total': total,
            'quantidade': qtd,
            'por_plataforma': por_plataforma,
            'por_campeonato': por_campeonato,
            'por_grupo': por_grupo,
            'dataframe': df_f
        }

        print(f"   Total: R$ {total:,.2f} | Qtd: {qtd:,}")

    def estimar_transporte(self):
        """Estima transporte baseado em distribuição proporcional de Passagens + Hospedagens"""
        print("\n🔮 Estimando TRANSPORTE baseado em Passagens + Hospedagens...")
        
        try:
            if 'passagens' not in self.dados_consolidados or 'hospedagens' not in self.dados_consolidados:
                print("   ⚠️  Dados de Passagens ou Hospedagens não disponíveis para estimativa")
                return False
            
            df_pass = self.dados_consolidados['passagens']['dataframe'].copy()
            df_hosp = self.dados_consolidados['hospedagens']['dataframe'].copy()
            
            if len(df_pass) == 0 and len(df_hosp) == 0:
                print("   ⚠️  Nenhum dado disponível para estimativa")
                return False
            
            if 'Data' in df_pass.columns:
                df_pass['Data'] = pd.to_datetime(df_pass['Data'], errors='coerce')
                df_pass['Mês'] = df_pass['Data'].dt.month
                df_pass['Ano'] = df_pass['Data'].dt.year
            else:
                df_pass['Mês'] = 1
                df_pass['Ano'] = datetime.now().year
            
            if 'Data' in df_hosp.columns:
                df_hosp['Data'] = pd.to_datetime(df_hosp['Data'], errors='coerce')
                df_hosp['Mês'] = df_hosp['Data'].dt.month
                df_hosp['Ano'] = df_hosp['Data'].dt.year
            else:
                df_hosp['Mês'] = 1
                df_hosp['Ano'] = datetime.now().year
            
            if len(df_pass) > 0 and 'VALOR_AJUSTADO' in df_pass.columns:
                df_pass_group = df_pass.groupby(['Mês', 'Ano', 'AREA', 'NOME_PROJETO'])['VALOR_AJUSTADO'].sum().reset_index()
                df_pass_group.columns = ['Mês', 'Ano', 'Área', 'Campeonato', 'Valor_Passagens']
            else:
                df_pass_group = pd.DataFrame(columns=['Mês', 'Ano', 'Área', 'Campeonato', 'Valor_Passagens'])
            
            if len(df_hosp) > 0 and 'TOTAL_AJUSTADO' in df_hosp.columns:
                df_hosp_group = df_hosp.groupby(['Mês', 'Ano', 'AREA', 'NOME_PROJETO'])['TOTAL_AJUSTADO'].sum().reset_index()
                df_hosp_group.columns = ['Mês', 'Ano', 'Área', 'Campeonato', 'Valor_Hospedagens']
            else:
                df_hosp_group = pd.DataFrame(columns=['Mês', 'Ano', 'Área', 'Campeonato', 'Valor_Hospedagens'])
            
            if len(df_pass_group) > 0 and len(df_hosp_group) > 0:
                df_base = pd.merge(df_pass_group, df_hosp_group, 
                                   on=['Mês', 'Ano', 'Área', 'Campeonato'], 
                                   how='outer').fillna(0)
            elif len(df_pass_group) > 0:
                df_base = df_pass_group.copy()
                df_base['Valor_Hospedagens'] = 0
            elif len(df_hosp_group) > 0:
                df_base = df_hosp_group.copy()
                df_base['Valor_Passagens'] = 0
            else:
                print("   ⚠️  Não foi possível criar base para estimativa")
                return False
            
            df_base['Valor_Total_Base'] = df_base.get('Valor_Passagens', 0) + df_base.get('Valor_Hospedagens', 0)
            total_base = df_base['Valor_Total_Base'].sum()
            
            if total_base == 0:
                print("   ⚠️  Total base zero, não é possível estimar")
                return False
            
            total_transporte_real = self.dados_consolidados['transporte']['total']
            
            if total_transporte_real > 0:
                fator_transporte = total_transporte_real / total_base
                print(f"   💡 Distribuindo transporte real (R$ {total_transporte_real:,.2f}) proporcionalmente")
            else:
                fator_transporte = 0.143
                total_transporte_estimado = total_base * fator_transporte
                print(f"   💡 Estimando transporte como {fator_transporte*100:.1f}% do total (R$ {total_transporte_estimado:,.2f})")
            
            df_base['Proporcao'] = df_base['Valor_Total_Base'] / total_base
            df_base['Transporte_Estimado'] = df_base['Proporcao'] * (total_transporte_real if total_transporte_real > 0 else total_base * fator_transporte)
            
            por_grupo_est = df_base.groupby('Área')['Transporte_Estimado'].agg(['sum', 'count']).reset_index()
            por_grupo_est.columns = ['Grupo', 'Valor', 'Quantidade']
            por_grupo_est = por_grupo_est.sort_values('Valor', ascending=False)
            
            por_campeonato_est = df_base.groupby('Campeonato')['Transporte_Estimado'].agg(['sum', 'count']).reset_index()
            por_campeonato_est.columns = ['Campeonato', 'Valor', 'Quantidade']
            por_campeonato_est = por_campeonato_est.sort_values('Valor', ascending=False)
            
            if len(df_pass) > 0 and 'PRODUTO_MACRO' in df_pass.columns:
                df_base_plat = df_base.merge(
                    df_pass[['NOME_PROJETO', 'PRODUTO_MACRO']].drop_duplicates(),
                    left_on='Campeonato',
                    right_on='NOME_PROJETO',
                    how='left'
                )
                por_plataforma_est = df_base_plat.groupby('PRODUTO_MACRO')['Transporte_Estimado'].agg(['sum', 'count']).reset_index()
                por_plataforma_est.columns = ['Plataforma', 'Valor', 'Quantidade']
                por_plataforma_est = por_plataforma_est.sort_values('Valor', ascending=False)
            else:
                por_plataforma_est = pd.DataFrame(columns=['Plataforma', 'Valor', 'Quantidade'])
            
            total_estimado = por_grupo_est['Valor'].sum()
            
            self.dados_consolidados['transporte_estimado'] = {
                'total': total_estimado,
                'quantidade': len(df_base),
                'por_plataforma': por_plataforma_est,
                'por_campeonato': por_campeonato_est,
                'por_grupo': por_grupo_est,
                'dataframe': df_base,
                'metodo': 'proporcional' if total_transporte_real > 0 else 'estimado_14.3%'
            }
            
            print(f"   ✅ Transporte estimado: R$ {total_estimado:,.2f} | Qtd: {len(df_base):,}")
            return True
            
        except Exception as e:
            print(f"   ⚠️  Erro ao estimar transporte: {e}")
            traceback.print_exc()
            return False

    def extrair_transporte(self):
        """Extrai dados de transporte (Uber/99 + Aluguel)"""
        print("\n🚗 Processando TRANSPORTE (Uber/99 + Aluguel)...")
        
        nome_aba = 'Consolidado Geral (UBER e 99)'
        
        aba_existe = False
        if self.wb and nome_aba in self.wb.sheetnames:
            aba_existe = True
        
        if not aba_existe:
            print(f"   ⚠️  Aba '{nome_aba}' não encontrada - será usado transporte estimado")
            self.dados_consolidados['transporte'] = {
                'total': 0,
                'quantidade': 0,
                'por_plataforma': pd.DataFrame(columns=['Plataforma', 'Valor', 'Quantidade']),
                'por_campeonato': pd.DataFrame(columns=['Campeonato', 'Valor', 'Quantidade']),
                'por_grupo': pd.DataFrame(columns=['Grupo', 'Valor', 'Quantidade']),
                'dataframe': pd.DataFrame()
            }
            print("   💡 Transporte será estimado automaticamente após processar Passagens e Hospedagens")
            return
        
        try:
            df = pd.read_excel(str(self.caminho_planilha), sheet_name=nome_aba)
        except Exception as e:
            raise Exception(f"❌ Erro ao ler aba {nome_aba}: {e}")
        
        # Remover linhas de subtotal (ÁREA == 'Total')
        df = df[df['ÁREA'].astype(str).str.strip() != 'Total']
        
        # Filtrar MÊS <= 10
        operador = self.filtros.get('OPERADOR_MES')
        mes = self.filtros.get('MES')
        if 'Mês' in df.columns and mes is not None and operador:
            if operador == "=":
                df = df[df['Mês'] == mes]
            elif operador == "<=":
                df = df[df['Mês'] <= mes]
        
        # ✅ FILTRO CRÍTICO: Apenas as áreas específicas por Base
        areas_uber = ['Gestão De Eventos', 'Multimodalidades', 'Futebol', 'Produção']
        areas_aluguel = ['Gestão De Eventos', 'Multimodalidades', 'Futebol', 'Produção', 'Direção', 'Grandes Eventos']
        
        df_uber = df[(df['Base'] == 'UBER/99') & (df['ÁREA'].isin(areas_uber))].copy()
        df_aluguel = df[(df['Base'] == 'ALUGUEL') & (df['ÁREA'].isin(areas_aluguel))].copy()
        
        # ✅ Descontar Programas de Gestão de Eventos (UBER/99)
        soma_programas_uber = df_uber[(df_uber['ÁREA'] == 'Gestão De Eventos') & (df_uber['Programas'] == 'x')]['Total'].sum()
        
        # Consolidar
        df = pd.concat([df_uber, df_aluguel], ignore_index=True)
        
        if len(df) == 0:
            print("   ⚠️  Nenhum dado válido encontrado em Transporte")
            total = 0
            qtd = 0
            por_grupo = pd.DataFrame(columns=['Grupo', 'Valor', 'Quantidade'])
        else:
            total = df['Total'].sum() - soma_programas_uber
            qtd = len(df)
            por_grupo = df.groupby('ÁREA')['Total'].agg(['sum', 'count']).reset_index()
            por_grupo.columns = ['Grupo', 'Valor', 'Quantidade']
            
            # Ajustar Gestão de Eventos
            if 'Gestão De Eventos' in por_grupo['Grupo'].values:
                idx = por_grupo[por_grupo['Grupo'] == 'Gestão De Eventos'].index[0]
                por_grupo.loc[idx, 'Valor'] -= soma_programas_uber
            
            por_grupo = por_grupo.sort_values('Valor', ascending=False)
        
        por_plataforma = pd.DataFrame(columns=['Plataforma', 'Valor', 'Quantidade'])
        por_campeonato = pd.DataFrame(columns=['Campeonato', 'Valor', 'Quantidade'])
        
        self.dados_consolidados['transporte'] = {
            'total': total,
            'quantidade': qtd,
            'por_plataforma': por_plataforma,
            'por_campeonato': por_campeonato,
            'por_grupo': por_grupo,
            'dataframe': df
        }
        
        print(f"   Total: R$ {total:,.2f} | Qtd: {qtd:,}")



    def _consolidar_dataframes(self, dfs, chave_agrupamento):
        """Método auxiliar para consolidar múltiplos DataFrames"""
        if not dfs:
            return pd.DataFrame(columns=[chave_agrupamento])
        
        dfs_validos = [df for df in dfs if len(df) > 0]
        
        if not dfs_validos:
            return pd.DataFrame(columns=[chave_agrupamento])
        
        resultado = dfs_validos[0].copy()
        
        for df in dfs_validos[1:]:
            resultado = pd.merge(resultado, df, on=chave_agrupamento, how='outer')
        
        return resultado.fillna(0)
    
    def consolidar_tudo(self):
        """Consolida todos os componentes"""
        print("\n📦 Consolidando TODOS os componentes...")
        
        total_transporte = self.dados_consolidados['transporte']['total']
        if total_transporte == 0 and 'transporte_estimado' in self.dados_consolidados:
            total_transporte = self.dados_consolidados['transporte_estimado']['total']
        
        total_geral = (
            self.dados_consolidados['passagens']['total'] +
            self.dados_consolidados['hospedagens']['total'] +
            total_transporte
        )
        
        qtd_transporte = self.dados_consolidados['transporte']['quantidade']
        if qtd_transporte == 0 and 'transporte_estimado' in self.dados_consolidados:
            qtd_transporte = self.dados_consolidados['transporte_estimado']['quantidade']
        
        qtd_total = (
            self.dados_consolidados['passagens']['quantidade'] +
            self.dados_consolidados['hospedagens']['quantidade'] +
            qtd_transporte
        )
        
        # Consolidar por Plataforma
        dfs_plat = []
        for componente in ['passagens', 'hospedagens']:
            df = self.dados_consolidados[componente]['por_plataforma'].copy()
            if len(df) > 0:
                df.columns = ['Plataforma', f'Valor_{componente}', f'Qtd_{componente}']
                dfs_plat.append(df)
        
        df_plat_consolidado = self._consolidar_dataframes(dfs_plat, 'Plataforma')
        if len(df_plat_consolidado) > 0:
            for col in ['Valor_passagens', 'Valor_hospedagens', 'Qtd_passagens', 'Qtd_hospedagens']:
                if col not in df_plat_consolidado.columns:
                    df_plat_consolidado[col] = 0
            
            df_plat_consolidado['Valor_Total'] = (
                df_plat_consolidado['Valor_passagens'] + 
                df_plat_consolidado['Valor_hospedagens']
            )
            df_plat_consolidado['Qtd_Total'] = (
                df_plat_consolidado['Qtd_passagens'] + 
                df_plat_consolidado['Qtd_hospedagens']
            )
            df_plat_consolidado = self._ordenar_plataformas(df_plat_consolidado)
        
        # Consolidar por Campeonato
        dfs_camp = []
        for componente in ['passagens', 'hospedagens']:
            df = self.dados_consolidados[componente]['por_campeonato'].copy()
            if len(df) > 0:
                df.columns = ['Campeonato', f'Valor_{componente}', f'Qtd_{componente}']
                dfs_camp.append(df)
        
        df_camp_consolidado = self._consolidar_dataframes(dfs_camp, 'Campeonato')
        if len(df_camp_consolidado) > 0:
            for col in ['Valor_passagens', 'Valor_hospedagens', 'Qtd_passagens', 'Qtd_hospedagens']:
                if col not in df_camp_consolidado.columns:
                    df_camp_consolidado[col] = 0
            
            df_camp_consolidado['Valor_Total'] = (
                df_camp_consolidado['Valor_passagens'] + 
                df_camp_consolidado['Valor_hospedagens']
            )
            df_camp_consolidado['Qtd_Total'] = (
                df_camp_consolidado['Qtd_passagens'] + 
                df_camp_consolidado['Qtd_hospedagens']
            )
            df_camp_consolidado = df_camp_consolidado.sort_values('Valor_Total', ascending=False)
        
        # Consolidar por Grupo
        dfs_grupo = []
        for componente in ['passagens', 'hospedagens', 'transporte']:
            df = self.dados_consolidados[componente]['por_grupo'].copy()
            if len(df) > 0:
                df.columns = ['Grupo', f'Valor_{componente}', f'Qtd_{componente}']
                dfs_grupo.append(df)
        
        if 'transporte_estimado' in self.dados_consolidados:
            df_trans_est = self.dados_consolidados['transporte_estimado']['por_grupo'].copy()
            if len(df_trans_est) > 0:
                df_trans_est.columns = ['Grupo', 'Valor_transporte_estimado', 'Qtd_transporte_estimado']
                dfs_grupo.append(df_trans_est)
        
        df_grupo_consolidado = self._consolidar_dataframes(dfs_grupo, 'Grupo')
        if len(df_grupo_consolidado) > 0:
            for col in ['Valor_passagens', 'Valor_hospedagens', 'Valor_transporte', 
                       'Qtd_passagens', 'Qtd_hospedagens', 'Qtd_transporte']:
                if col not in df_grupo_consolidado.columns:
                    df_grupo_consolidado[col] = 0
            
            valor_transporte = df_grupo_consolidado['Valor_transporte'].fillna(0)
            if 'Valor_transporte_estimado' in df_grupo_consolidado.columns:
                valor_transporte = valor_transporte + df_grupo_consolidado['Valor_transporte_estimado'].fillna(0)
            
            df_grupo_consolidado['Valor_Total'] = (
                df_grupo_consolidado['Valor_passagens'].fillna(0) + 
                df_grupo_consolidado['Valor_hospedagens'].fillna(0) +
                valor_transporte
            )
            
            qtd_transporte = df_grupo_consolidado['Qtd_transporte'].fillna(0)
            if 'Qtd_transporte_estimado' in df_grupo_consolidado.columns:
                qtd_transporte = qtd_transporte + df_grupo_consolidado['Qtd_transporte_estimado'].fillna(0)
            
            df_grupo_consolidado['Qtd_Total'] = (
                df_grupo_consolidado['Qtd_passagens'].fillna(0) + 
                df_grupo_consolidado['Qtd_hospedagens'].fillna(0) +
                qtd_transporte
            )
            df_grupo_consolidado = df_grupo_consolidado.sort_values('Valor_Total', ascending=False)
        
        self.dados_consolidados['consolidado'] = {
            'total_geral': total_geral,
            'quantidade_total': qtd_total,
            'por_plataforma': df_plat_consolidado,
            'por_campeonato': df_camp_consolidado,
            'por_grupo': df_grupo_consolidado
        }
        
        if total_geral > 0:
            pct_passagens = (self.dados_consolidados['passagens']['total'] / total_geral) * 100
            pct_hospedagens = (self.dados_consolidados['hospedagens']['total'] / total_geral) * 100
            pct_transporte = (total_transporte / total_geral) * 100
        else:
            pct_passagens = pct_hospedagens = pct_transporte = 0
        
        print(f"   Total Geral: R$ {total_geral:,.2f}")
        print(f"   Passagens: R$ {self.dados_consolidados['passagens']['total']:,.2f} ({pct_passagens:.1f}%)")
        print(f"   Hospedagens: R$ {self.dados_consolidados['hospedagens']['total']:,.2f} ({pct_hospedagens:.1f}%)")
        if self.dados_consolidados['transporte']['total'] > 0:
            print(f"   Transporte: R$ {self.dados_consolidados['transporte']['total']:,.2f} ({pct_transporte:.1f}%)")
        elif 'transporte_estimado' in self.dados_consolidados:
            print(f"   Transporte (ESTIMADO): R$ {total_transporte:,.2f} ({pct_transporte:.1f}%)")
        else:
            print(f"   Transporte: R$ {total_transporte:,.2f} ({pct_transporte:.1f}%)")
        
        self.dados_consolidados['consolidado']['percentuais'] = {
            'passagens': pct_passagens,
            'hospedagens': pct_hospedagens,
            'transporte': pct_transporte
        }
    
    def gerar_planilha_tratada(self, caminho_saida):
        """Gera planilha Excel consolidada e tratada"""
        print(f"\n💾 Gerando planilha tratada em {caminho_saida}...")
        caminho_saida = Path(caminho_saida)
        caminho_saida.parent.mkdir(parents=True, exist_ok=True)
        
        pct = self.dados_consolidados['consolidado'].get('percentuais', {})
        pct_passagens = pct.get('passagens', 0)
        pct_hospedagens = pct.get('hospedagens', 0)
        pct_transporte = pct.get('transporte', 0)
        
        with pd.ExcelWriter(str(caminho_saida), engine='openpyxl') as writer:
            total_transporte = self.dados_consolidados['transporte']['total']
            qtd_transporte = self.dados_consolidados['transporte']['quantidade']
            transporte_label = 'Transporte (Uber/99)'
            if total_transporte == 0 and 'transporte_estimado' in self.dados_consolidados:
                total_transporte = self.dados_consolidados['transporte_estimado']['total']
                qtd_transporte = self.dados_consolidados['transporte_estimado']['quantidade']
                transporte_label = 'Transporte (Uber/99) - ESTIMADO'
            
            resumo_data = {
                'Componente': ['Passagens', 'Hospedagens', transporte_label, 'TOTAL'],
                'Valor (R$)': [
                    self.dados_consolidados['passagens']['total'],
                    self.dados_consolidados['hospedagens']['total'],
                    total_transporte,
                    self.dados_consolidados['consolidado']['total_geral']
                ],
                'Quantidade': [
                    self.dados_consolidados['passagens']['quantidade'],
                    self.dados_consolidados['hospedagens']['quantidade'],
                    qtd_transporte,
                    self.dados_consolidados['consolidado']['quantidade_total']
                ],
                'Percentual (%)': [
                    pct_passagens, pct_hospedagens, pct_transporte, 100.0
                ]
            }
            df_resumo = pd.DataFrame(resumo_data)
            df_resumo.to_excel(writer, sheet_name='Resumo Executivo', index=False)
            
            df = self.dados_consolidados['consolidado']['por_plataforma']
            df.to_excel(writer, sheet_name='Por Plataforma (MACRO)', index=False)
            df = self.dados_consolidados['consolidado']['por_campeonato']
            df.to_excel(writer, sheet_name='Por Campeonato (MICRO)', index=False)
            df = self.dados_consolidados['consolidado']['por_grupo']
            df.to_excel(writer, sheet_name='Por Grupo de Pessoas', index=False)
            
            self.dados_consolidados['passagens']['por_plataforma'].to_excel(writer, sheet_name='Passagens - Plataforma', index=False)
            self.dados_consolidados['passagens']['por_campeonato'].head(100).to_excel(writer, sheet_name='Passagens - Campeonato', index=False)
            self.dados_consolidados['passagens']['por_grupo'].to_excel(writer, sheet_name='Passagens - Grupo', index=False)
            
            self.dados_consolidados['hospedagens']['por_plataforma'].to_excel(writer, sheet_name='Hospedagens - Plataforma', index=False)
            self.dados_consolidados['hospedagens']['por_campeonato'].head(100).to_excel(writer, sheet_name='Hospedagens - Campeonato', index=False)
            self.dados_consolidados['hospedagens']['por_grupo'].to_excel(writer, sheet_name='Hospedagens - Grupo', index=False)
            
            if len(self.dados_consolidados['transporte']['por_grupo']) > 0:
                self.dados_consolidados['transporte']['por_grupo'].to_excel(writer, sheet_name='Transporte - Grupo', index=False)
            
            if 'transporte_estimado' in self.dados_consolidados and len(self.dados_consolidados['transporte_estimado']['por_grupo']) > 0:
                self.dados_consolidados['transporte_estimado']['por_grupo'].to_excel(writer, sheet_name='Transporte Estimado - Grupo', index=False)
            
            if 'transporte_estimado' in self.dados_consolidados and len(self.dados_consolidados['transporte_estimado']['por_campeonato']) > 0:
                self.dados_consolidados['transporte_estimado']['por_campeonato'].head(100).to_excel(writer, sheet_name='Transporte Estimado - Campeonato', index=False)
        
        print(f"✅ Planilha tratada gerada com sucesso em: {caminho_saida}")
        
    def analisar_com_ia(self):
        """Usa IA para análise inteligente dos dados"""
        if not self.client:
            print("\n⚠️  Análise com IA não disponível")
            return None
        
        print("\n🤖 Executando análise com IA...")
        
        resumo = f"""
Análise do Painel Contábil - Dados Consolidados:

TOTAL GERAL DE LOGÍSTICA: R$ {self.dados_consolidados['consolidado']['total_geral']:,.2f}

Componentes:
- Passagens: R$ {self.dados_consolidados['passagens']['total']:,.2f}
- Hospedagens: R$ {self.dados_consolidados['hospedagens']['total']:,.2f}
- Transporte: R$ {self.dados_consolidados['transporte']['total']:,.2f}

Top 5 Plataformas (MACRO):
{self.dados_consolidados['consolidado']['por_plataforma'].head(5)[['Plataforma', 'Valor_Total']].to_string(index=False)}

Top 5 Campeonatos (MICRO):
{self.dados_consolidados['consolidado']['por_campeonato'].head(5)[['Campeonato', 'Valor_Total']].to_string(index=False)}

Top 5 Grupos de Pessoas:
{self.dados_consolidados['consolidado']['por_grupo'].head(5)[['Grupo', 'Valor_Total']].to_string(index=False)}
"""
        
        try:
            model = "gpt-4o-mini"
            
            response = self.client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "Você é um analista financeiro especializado em análise de dados contábeis e orçamentários. Forneça insights práticos e acionáveis."},
                    {"role": "user", "content": f"Analise os seguintes dados e forneça: (1) principais insights, (2) possíveis áreas de atenção, (3) recomendações para otimização:\n\n{resumo}"}
                ],
                max_tokens=1000,
                temperature=0.7
            )
            
            analise_ia = response.choices[0].message.content
            print("\n" + "="*80)
            print("ANÁLISE INTELIGENTE (IA)")
            print("="*80)
            print(analise_ia)
            print("="*80)
            
            self.dados_consolidados['analise_ia'] = analise_ia
            return analise_ia
            
        except Exception as e:
            print(f"⚠️  Erro na análise com IA: {e}")
            print(f"   Detalhes: {traceback.format_exc()}")
            return None
    
    def extrair_orcados(self):
        """
        Extrai valores orçados da aba RESUMO LOGÍSTICA
        Retorna dict com valores orçados por categoria
        """
        try:
            print("\n💰 Extraindo valores orçados...")
            df_orcado = pd.read_excel(str(self.caminho_planilha), sheet_name='RESUMO LOGÍSTICA', header=None)
            
            orcados = {
                'total': float(df_orcado.iloc[6, 1]),
                'passagens': float(df_orcado.iloc[10, 1]),
                'hospedagens': float(df_orcado.iloc[14, 1]),
                'transporte': float(df_orcado.iloc[18, 1])
            }
            
            print("✅ Valores orçados extraídos com sucesso!")
            print(f"   Total Orçado: R$ {orcados['total']:,.2f}")
            print(f"   Passagens Orçado: R$ {orcados['passagens']:,.2f}")
            print(f"   Hospedagens Orçado: R$ {orcados['hospedagens']:,.2f}")
            print(f"   Transporte Orçado: R$ {orcados['transporte']:,.2f}")
            
            # Armazenar nos dados consolidados
            self.dados_consolidados['orcados'] = orcados
            
            return orcados
            
        except Exception as e:
            print(f"⚠️  Erro ao extrair orçados: {e}")
            print("   A aba 'RESUMO LOGÍSTICA' pode não existir ou ter estrutura diferente")
            return None
    
    def salvar_analise_ia(self, caminho_arquivo):
        if 'analise_ia' not in self.dados_consolidados or not self.dados_consolidados['analise_ia']:
            print("⚠️  Nenhuma análise de IA disponível para salvar")
            return False
        try:
            caminho_arquivo = Path(caminho_arquivo)
            caminho_arquivo.parent.mkdir(parents=True, exist_ok=True)
            with open(caminho_arquivo, 'w', encoding='utf-8') as f:
                f.write("="*80 + "\n")
                f.write("ANÁLISE INTELIGENTE (IA) - PAINEL CONTÁBIL\n")
                f.write("="*80 + "\n")
                f.write(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                f.write("="*80 + "\n\n")
                f.write(self.dados_consolidados['analise_ia'])
                f.write("\n\n" + "="*80 + "\n")
            print(f"✅ Análise de IA salva em: {caminho_arquivo}")
            return True
        except Exception as e:
            print(f"⚠️  Erro ao salvar análise de IA: {e}")
            return False
        
    def executar_analise_completa(self, caminho_planilha_tratada, caminho_analise_ia=None):
        print("="*80)
        print("🚀 ANÁLISE COMPLETA V3 - COM IA, PLANILHA TRATADA E ORÇADO VS REALIZADO")
        print("="*80)
        
        inicio = datetime.now()
        
        try:
            self.carregar_planilha()
            self.extrair_passagens()
            self.extrair_hospedagens()
            self.extrair_transporte()
            
            if self.dados_consolidados['transporte']['total'] == 0:
                print("\n💡 Nenhum dado de transporte encontrado - gerando estimativa...")
                self.estimar_transporte()
            
            self.consolidar_tudo()
            self.gerar_planilha_tratada(caminho_planilha_tratada)
            
            # Extrair orcados
            orcados = self.extrair_orcados()
            
            analise_ia = self.analisar_com_ia()
            
            if caminho_analise_ia and analise_ia:
                self.salvar_analise_ia(caminho_analise_ia)
            
        except Exception as e:
            print(f"\n❌ ERRO durante a análise: {e}")
            print(traceback.format_exc())
            raise
        finally:
            if self.wb:
                try:
                    self.wb.close()
                except:
                    pass
        
        fim = datetime.now()
        duracao = (fim - inicio).total_seconds()
        
        print("\n" + "="*80)
        print(f"✅ ANÁLISE COMPLETA FINALIZADA EM {duracao:.2f}s")
        print("="*80)
        
        return {
            'total_geral': self.dados_consolidados['consolidado']['total_geral'],
            'passagens': self.dados_consolidados['passagens']['total'],
            'hospedagens': self.dados_consolidados['hospedagens']['total'],
            'transporte': self.dados_consolidados['transporte']['total'],
            'orcados': orcados,
            'analise_ia': analise_ia,
            'duracao': duracao
        }


def main():
    if os.name == 'nt':
        caminho_planilha = Path(r"C:\Users\ligomes\Downloads\painel_contabil_solucao_final\painel_contabil\Painel Contábil V2 - copia.xlsx")
        caminho_planilha_tratada = Path(__file__).parent / 'planilha_consolidada_tratada.xlsx'
        caminho_analise_ia = Path(__file__).parent / 'analise_ia.txt'
    else:
        caminho_planilha = Path('/home/ubuntu/upload/PainelContábilV2-copia.xlsx')
        caminho_planilha_tratada = Path('/home/ubuntu/painel_contabil/planilha_consolidada_tratada.xlsx')
        caminho_analise_ia = Path('/home/ubuntu/painel_contabil/analise_ia.txt')
    
    if not caminho_planilha.exists():
        print(f"❌ ERRO: Planilha não encontrada em {caminho_planilha}")
        print("💡 Por favor, ajuste o caminho no código ou coloque a planilha no local correto")
        return 1
    
    try:
        analisador = AnalisadorCompletoV3(caminho_planilha)
        resultado = analisador.executar_analise_completa(
            caminho_planilha_tratada, 
            caminho_analise_ia
        )
        
        print("\n📊 RESUMO FINAL:")
        print(f"  Total Geral: R$ {resultado['total_geral']:,.2f}")
        print(f"  Passagens: R$ {resultado['passagens']:,.2f}")
        print(f"  Hospedagens: R$ {resultado['hospedagens']:,.2f}")
        print(f"  Transporte: R$ {resultado['transporte']:,.2f}")
        print(f"  Planilha Tratada: {caminho_planilha_tratada}")
        if resultado['analise_ia']:
            print(f"  Análise IA: {caminho_analise_ia}")
        print(f"  Tempo de execução: {resultado['duracao']:.2f}s")
        
        return 0
    except Exception as e:
        print(f"\n❌ ERRO na execução: {e}")
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    main()